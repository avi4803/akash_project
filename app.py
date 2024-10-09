from flask import Flask, request, jsonify, render_template, send_file, send_from_directory
import cohere  # Ensure you have the Cohere API library installed
import os
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from PyPDF2 import PageObject, PdfReader, PdfWriter
from io import BytesIO
import uuid
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
app = Flask(__name__)

# Initialize the Cohere client
api_key = '3v1gSXcJA8TXjPUHe0kPcCyzEMY7Qo5f52M82SuN'
co = cohere.Client(api_key)


def generate_unique_filename(base_name):
    # Generate a 2-digit random number
    random_number = str(uuid.uuid4().int)[:2]  # Take the first 2 digits from UUID
    
    # Get the current timestamp
    timestamp = datetime.now().strftime('%Y%m%d%H%M')
    
    # Combine base name, random number, and timestamp to form a unique filename
    unique_filename = f"{base_name}_{random_number}_{timestamp}.pdf"
    return unique_filename
# Define a function to read prompt from file and format it with user data
def load_and_format_prompt(prompt_name, data):
    file_path = os.path.join('prompts', f'{prompt_name}_prompt.txt')
    excel_path = "userData.xlsx"

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Prompt file for {prompt_name} not found.")
    
    # Read the prompt template
    with open(file_path, 'r') as file:
        prompt_template = file.read()
    
    # Ensure that all keys in data have corresponding placeholders in the template
    try:
        prompt = prompt_template.format(**data)
    except KeyError as e:
        raise KeyError(f"Missing key in template: {e}")

    # Get current date and time
    current_date = datetime.now().date()
    current_time = datetime.now().time()

    # Load or create the Excel workbook
    if os.path.exists(excel_path):
        workbook = openpyxl.load_workbook(excel_path)
    else:
        workbook = openpyxl.Workbook()
        # Remove the default sheet created
        if 'Sheet1' in workbook.sheetnames:
            default_sheet = workbook['Sheet1']
            workbook.remove(default_sheet)
    
    # Select or create a sheet based on the prompt name
    if prompt_name in workbook.sheetnames:
        sheet = workbook[prompt_name]
    else:
        sheet = workbook.create_sheet(title=prompt_name)

    # Write the data to the selected sheet according to column headers
    if sheet.max_row == 1 and sheet.max_column == 1:  # Check if the sheet is new and empty
        # Add headers
        headers = list(data.keys()) + ['Date', 'Time']
        sheet.append(headers)
    else:
        headers = [cell.value for cell in sheet[1]]
    # Prepare the row data
    row_data = [data.get(header, '') for header in headers if header in data]
    row_data += [current_date, current_time]

    # Append the row data to the sheet
    sheet.append(row_data)

    # Save the workbook
    workbook.save(excel_path)
    return prompt

@app.route('/')
def index():
    return render_template('index.html')
# Define a function to generate PDF
  
def add_content_to_pdf(template_path, content, output_path,title):
    temp_pdf_path = "temp_content.pdf"
    
    doc = SimpleDocTemplate(temp_pdf_path, pagesize=letter)
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        name='Title',
        fontSize=18,
        leading=22,
        alignment=1,  # Center alignment
        textColor=colors.HexColor("#FF0000")  # Red color
    )
    
    heading_style = ParagraphStyle(
        name='Heading',
        fontSize=14,
        leading=18,
        spaceAfter=10,
        textColor=colors.HexColor("#007bff")
    )

    body_style = ParagraphStyle(
        name='Body',
        fontSize=12,
        leading=14,
        spaceAfter=10,
        textColor=colors.HexColor("#333333")
    )

    content_list = []
    
    content_list.append(Paragraph(title, title_style))
    content_list.append(Spacer(1, 20))
    
    paragraphs = content.split('\n')
    for para in paragraphs:
        para = para.strip()
        
        # Replace `**text**` with `<b>text</b>` for bold formatting
        while '**' in para:
            start_idx = para.index('**')
            end_idx = para.index('**', start_idx + 2)
            bold_text = para[start_idx + 2:end_idx]
            para = para[:start_idx] + '<b>' + bold_text + '</b>' + para[end_idx + 2:]
        
        # Handle content formatting
        if para.startswith("# "):
            content_list.append(Paragraph(para.replace("# ", ""), heading_style))
        elif para.startswith("## "):
            content_list.append(Paragraph(para.replace("## ", ""), heading_style))
        else:
            # Convert `- **text**` to `- <b>text</b>`
            para = para.replace('- <b>', '- <b>').replace('</b>', '</b>')
            content_list.append(Paragraph(para, body_style))
        
        content_list.append(Spacer(1, 12))
    
    doc.build(content_list)
    
    # Here would be the code to merge the content with the template PDF, 
    # which is not repeated to focus on the bold text formatting.

    template_reader = PdfReader(template_path)
    content_reader = PdfReader(temp_pdf_path)
    writer = PdfWriter()

    background_page = template_reader.pages[0]

    for page_num in range(len(content_reader.pages)):
        new_page = PageObject.create_blank_page(width=background_page.mediabox.width, height=background_page.mediabox.height)
        new_page.merge_page(background_page)
        
        content_page = content_reader.pages[page_num]
        new_page.merge_page(content_page)

        writer.add_page(new_page)

    with open(output_path, 'wb') as output_pdf:
        writer.write(output_pdf)

    os.remove(temp_pdf_path)
    print(f"PDF successfully saved at: {output_path}")

@app.route('/business-growth', methods=['GET', 'POST'])
def business_growth():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('business_growth', data)
        response = co.generate(prompt=prompt, 
                               model='command-r-plus', 
                               temperature=0.5)
        strategy = response.generations[0].text
        # output_folder = 'pdfs'
        # os.makedirs(output_folder, exist_ok=True)
        # output_path = os.path.join(output_folder, 'generated_business_growth_strategy.pdf')
        pdf_filename = 'business_growth_strategy.pdf  '
        unique_filename = generate_unique_filename(pdf_filename)
        # Save the PDF to a file
        pdf_path = os.path.join('pdfs', unique_filename)
        title="Business Growth Strategy"
        add_content_to_pdf('template.pdf', strategy, pdf_path,title)

        return jsonify({'strategy': strategy, 'pdf_filename': unique_filename})
    return render_template('business_growth.html')

@app.route('/lead-generation', methods=['GET', 'POST'])
def lead_generation():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('lead_gen', data)
        response = co.generate(prompt=prompt,
                                model='command-r-plus', 
                                temperature=0.5)
        strategy = response.generations[0].text
        
        pdf_filename = 'lead_generation.pdf'

        unique_filename = generate_unique_filename(pdf_filename)
        # Save the PDF to a file
        pdf_path = os.path.join('pdfs', unique_filename)
        title="Lead Generation Strategy"
        add_content_to_pdf('template.pdf', strategy, pdf_path,title)

        return jsonify({'strategy': strategy, 'pdf_filename': unique_filename})
    return render_template('lead_generation.html')

@app.route('/funding-pitch', methods=['GET', 'POST'])
def funding_pitch():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('funding_pitch', data)
        response = co.generate(prompt=prompt, 
                               model='command-r-plus', 
                               temperature=0.5)
        strategy = response.generations[0].text
       
        pdf_filename = 'funding_pitch.pdf'

        # Save the PDF to a file
        unique_filename = generate_unique_filename(pdf_filename)
        # Save the PDF to a file
        pdf_path = os.path.join('pdfs', unique_filename)
        title="Funding Pitch Strategy"
        add_content_to_pdf('template.pdf', strategy, pdf_path,title)

        return jsonify({'strategy': strategy, 'pdf_filename':unique_filename})
    return render_template('funding_pitch.html')

@app.route('/social-media-strategy', methods=['GET', 'POST'])
def social_media_strategy():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('social_media_strategy', data)
        response = co.generate(prompt=prompt,
                                model='command-r-plus', 
                                temperature=0.5)
        strategy = response.generations[0].text
       
        pdf_filename = 'social_media_strategy.pdf'
        unique_filename = generate_unique_filename(pdf_filename)
        # Save the PDF to a file
        pdf_path = os.path.join('pdfs', unique_filename)
        title="Social Media Strategy"
        add_content_to_pdf('template.pdf', strategy, pdf_path,title)

        return jsonify({'strategy': strategy, 'pdf_filename': unique_filename})
    return render_template('social_media_strategy.html')

@app.route('/business-queries', methods=['GET', 'POST'])
def business_queries():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('business_queries', data)
        response = co.generate(prompt=prompt,
                                model='command-r-plus',
                                  temperature=0.5)
        strategy = response.generations[0].text
        
        pdf_filename = 'business_queries.pdf'
        unique_filename = generate_unique_filename(pdf_filename)
        # Save the PDF to a file
        pdf_path = os.path.join('pdfs', unique_filename)
        title="Business Query Strategy"
        add_content_to_pdf('template.pdf', strategy, pdf_path,title)

        return jsonify({'strategy': strategy, 'pdf_filename': unique_filename})
    return render_template('business_queries.html')

@app.route('/linkedin-strategy', methods=['GET', 'POST'])
def linkedin_strategy():
    if request.method == 'POST':
        data = request.json
        prompt = load_and_format_prompt('linkedin_strategy', data)
        response = co.generate(prompt=prompt,
                                model='command-r-plus',
                                  temperature=0.5)
        strategy = response.generations[0].text
        
        pdf_filename = 'linkedin_strategy.pdf'

        unique_filename = generate_unique_filename(pdf_filename)
        # Save the PDF to a file
        pdf_path = os.path.join('pdfs', unique_filename)
        title="Linkedin Strategy"
        add_content_to_pdf('template.pdf', strategy, pdf_path,title)

        return jsonify({'strategy': strategy, 'pdf_filename': unique_filename})
    return render_template('linkedin_strategy.html')
@app.route('/pdfs/<filename>')
def download_pdf(filename):
    print(f"Serving file from directory 'pdfs' with filename: {filename}")
    
    # Check if the file exist
    
    return send_file(os.path.join('pdfs', filename), as_attachment=True)

if __name__ == '__main__':
    if not os.path.exists('pdfs'):
        os.makedirs('pdfs')
    app.run(debug=True)
